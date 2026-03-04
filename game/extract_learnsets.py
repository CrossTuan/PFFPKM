import json
from pathlib import Path
from openpyxl import load_workbook

# Path to the spreadsheet and output JSON
spreadsheet = Path(__file__).parent.parent.parent / "Multiversal Movepool v2.1.0 (8-6-24).xlsx"
output = Path(__file__).parent / "learnsets.json"


wb = load_workbook(spreadsheet, data_only=True)
learnsets = {}

# Find the 'SortedMoveset' sheet
sheet = None
for ws in wb.worksheets:
    if ws.title.lower() == 'sortedmoveset':
        sheet = ws
        break
if not sheet:
    raise Exception("Could not find 'SortedMoveset' sheet in the workbook.")



# Parse rows: [#, Name, Lvl1, Move1, Lvl2, Move2, ...]
species_moves = {}
for row in sheet.iter_rows(min_row=2):
    name = row[1].value
    if not name:
        continue
    species_key = str(name).strip().lower()
    if species_key not in species_moves:
        species_moves[species_key] = []
    # Iterate over (level, move) pairs: (C,D), (E,F), (G,H), ...
    for i in range(2, len(row)-1, 2):
        level = row[i].value
        move = row[i+1].value
        if isinstance(level, (int, float)) and move:
            species_moves[species_key].append((int(level), str(move).strip()))

# For each Pokémon, sort moves by level and remove duplicates (keep first occurrence)
for species_key, moves in species_moves.items():
    if species_key == "cubone":
        print(f"All moves for Cubone from SortedMoveset:")
        for lvl, mv in sorted(moves, key=lambda x: x[0]):
            print(f"  Level {lvl}: {mv}")
    seen = set()
    sorted_moves = []
    for lvl, mv in sorted(moves, key=lambda x: x[0]):
        if mv.lower() not in seen:
            sorted_moves.append({"level": lvl, "move": mv})
            seen.add(mv.lower())
    learnsets[species_key] = {"9": sorted_moves}

with open(output, "w", encoding="utf-8") as f:
    json.dump(learnsets, f, ensure_ascii=False, indent=2)

print(f"Extracted {len(learnsets)} Pokémon learnsets to {output}")

with open(output, "w", encoding="utf-8") as f:
    json.dump(learnsets, f, ensure_ascii=False, indent=2)

print(f"Extracted {len(learnsets)} Pokémon learnsets to {output}")
